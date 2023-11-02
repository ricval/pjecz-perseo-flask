"""
CLI Nominas
"""
import os
import re
from datetime import datetime
from pathlib import Path

import click
import xlrd
from dotenv import load_dotenv
from openpyxl import Workbook

from lib.safe_string import QUINCENA_REGEXP, safe_clave, safe_quincena, safe_string
from perseo.app import create_app
from perseo.blueprints.bancos.models import Banco
from perseo.blueprints.centros_trabajos.models import CentroTrabajo
from perseo.blueprints.conceptos.models import Concepto
from perseo.blueprints.conceptos_productos.models import ConceptoProducto
from perseo.blueprints.nominas.models import Nomina
from perseo.blueprints.percepciones_deducciones.models import PercepcionDeduccion
from perseo.blueprints.personas.models import Persona
from perseo.blueprints.plazas.models import Plaza
from perseo.blueprints.productos.models import Producto
from perseo.blueprints.quincenas.models import Quincena
from perseo.extensions import database

EXPLOTACION_BASE_DIR = os.environ.get("EXPLOTACION_BASE_DIR")
NOMINAS_FILENAME_XLS = "NominaFmt2.XLS"

app = create_app()
app.app_context().push()
database.app = app


@click.group()
def cli():
    """Nominas"""


@click.command()
@click.argument("quincena", type=str)
def alimentar(quincena: str):
    """Alimentar nominas"""

    # Iniciar sesion con la base de datos para que la alimentacion sea rapida
    sesion = database.session

    # Validar quincena
    if re.match(QUINCENA_REGEXP, quincena) is None:
        click.echo("Quincena inválida")
        return

    # Validar el directorio donde espera encontrar los archivos de explotacion
    if EXPLOTACION_BASE_DIR is None:
        click.echo("Variable de entorno EXPLOTACION_BASE_DIR no definida.")
        return

    # Validar si existe el archivo
    ruta = Path(EXPLOTACION_BASE_DIR, quincena, NOMINAS_FILENAME_XLS)
    if not ruta.exists():
        click.echo(f"AVISO: {str(ruta)} no se encontró.")
        return
    if not ruta.is_file():
        click.echo(f"AVISO: {str(ruta)} no es un archivo.")
        return

    # Revisar si existe el registro en quincenas, de lo contrario insertarlo
    quincena_obj = Quincena.query.filter_by(quincena=quincena).first()
    if quincena_obj is None:
        quincena_obj = Quincena(quincena=quincena, estado=Quincena.ESTADOS["ABIERTA"])
        sesion.add(quincena_obj)
        click.echo(f"  Quincena {quincena} insertada")

    # Abrir el archivo XLS con xlrd
    libro = xlrd.open_workbook(str(ruta))

    # Obtener la primera hoja
    hoja = libro.sheet_by_index(0)

    # Iniciar contador de percepciones-deducciones alimentadas
    contador = 0

    # Bucle por cada fila
    click.echo("Alimentando nominas...")
    for fila in range(1, hoja.nrows):
        # Tomar las columnas
        centro_trabajo_clave = hoja.cell_value(fila, 1)
        rfc = hoja.cell_value(fila, 2)
        nombre_completo = hoja.cell_value(fila, 3)
        plaza_clave = hoja.cell_value(fila, 8)
        percepcion = int(hoja.cell_value(fila, 12)) / 100.0
        deduccion = int(hoja.cell_value(fila, 13)) / 100.0
        impte = int(hoja.cell_value(fila, 14)) / 100.0
        modelo = int(hoja.cell_value(fila, 236))
        num_empleado = int(hoja.cell_value(fila, 240))

        # Separar nombre_completo, en apellido_primero, apellido_segundo y nombres
        separado = safe_string(nombre_completo, save_enie=True).split(" ")
        apellido_primero = separado[0]
        apellido_segundo = separado[1]
        nombres = " ".join(separado[2:])

        # Revisar si el Centro de Trabajo existe, de lo contrario insertarlo
        centro_trabajo = CentroTrabajo.query.filter_by(clave=centro_trabajo_clave).first()
        if centro_trabajo is None:
            centro_trabajo = CentroTrabajo(clave=centro_trabajo_clave, descripcion="ND")
            sesion.add(centro_trabajo)
            click.echo(f"  Centro de Trabajo {centro_trabajo_clave} insertado")

        # Revisar si la Persona existe, de lo contrario insertarlo
        persona = Persona.query.filter_by(rfc=rfc).first()
        if persona is None:
            persona = Persona(
                rfc=rfc,
                nombres=nombres,
                apellido_primero=apellido_primero,
                apellido_segundo=apellido_segundo,
                modelo=modelo,
                num_empleado=num_empleado,
            )
            sesion.add(persona)
            click.echo(f"  Persona {rfc} insertada")

        # Revisar si la Plaza existe, de lo contrario insertarla
        plaza = Plaza.query.filter_by(clave=plaza_clave).first()
        if plaza is None:
            plaza = Plaza(clave=plaza_clave, descripcion="ND")
            sesion.add(plaza)
            click.echo(f"  Plaza {plaza_clave} insertada")

        # Bucle entre P-D para determinar el tipo entre SALARIO y DESPENSA
        nomina_tipo = None
        col_num = 26
        while True:
            # Tomar el tipo y el conc para armar la clave del concepto
            tipo = safe_string(hoja.cell_value(fila, col_num))
            conc = safe_string(hoja.cell_value(fila, col_num + 1))
            concepto_clave = f"{tipo}{conc}"

            # Si el tipo es un texto vacio, se rompe el ciclo
            if tipo == "":
                break

            # Si el concepto_clave es PME, entonces es DESPENSA y se termina este ciclo
            if concepto_clave == "PME":
                nomina_tipo = Nomina.TIPOS["DESPENSA"]
                break

            # Incrementar col_num en SEIS
            col_num += 6

            # Romper el ciclo cuando se llega a la columna
            if col_num > 236:
                break

        # Si no se encontro el tipo, entonces es SALARIO
        if nomina_tipo is None:
            nomina_tipo = Nomina.TIPOS["SALARIO"]

        # Alimentar nomina
        nomina = Nomina(
            centro_trabajo=centro_trabajo,
            persona=persona,
            plaza=plaza,
            quincena=quincena,
            percepcion=percepcion,
            deduccion=deduccion,
            importe=impte,
            tipo=nomina_tipo,
        )
        sesion.add(nomina)

        # Incrementar contador
        contador += 1
        if contador % 100 == 0:
            click.echo(f"  Van {contador}...")

    # Cerrar la sesion para que se guarden todos los datos en la base de datos
    sesion.commit()
    sesion.close()

    # Mensaje termino
    click.echo(f"Nominas terminado: {contador} nominas alimentadas en la quincena {quincena}.")


@click.command()
@click.argument("quincena", type=str)
def generar_nominas(quincena: str):
    """Generar archivo XLSX con las nominas de una quincena"""

    # Validar quincena
    if re.match(QUINCENA_REGEXP, quincena) is None:
        click.echo("Quincena inválida.")
        return

    # Iniciar sesion con la base de datos para que la alimentacion sea rapida
    sesion = database.session

    # Cargar todos los bancos
    bancos = Banco.query.filter_by(estatus="A").all()

    # Bucle para igualar el consecutivo_generado al consecutivo
    # for banco in bancos:
    #     banco.consecutivo_generado = banco.consecutivo

    # Consultar las nominas de la quincena, solo tipo SALARIO
    nominas = Nomina.query.filter_by(quincena=quincena).filter_by(tipo="SALARIO").filter_by(estatus="A").all()

    # Iniciar el archivo XLSX
    libro = Workbook()

    # Tomar la hoja del libro XLSX
    hoja = libro.active

    # Agregar la fila con las cabeceras de las columnas
    hoja.append(
        [
            "QUINCENA",
            "CENTRO DE TRABAJO",
            "RFC",
            "NOMBRE COMPLETO",
            "NUMERO DE EMPLEADO",
            "MODELO",
            "PLAZA",
            "NOMBRE DEL BANCO",
            "BANCO ADMINISTRADOR",
            "NUMERO DE CUENTA",
            "MONTO A DEPOSITAR",
            "NO DE CHEQUE",
        ]
    )

    # Bucle para crear cada fila del archivo XLSX
    contador = 0
    personas_sin_cuentas = []
    for nomina in nominas:
        # Si el modelo de la persona es 3, se omite porque otro generar se encarga de ellos
        if nomina.persona.modelo == 3:
            continue

        # Tomar las cuentas de la persona
        cuentas = nomina.persona.cuentas

        # Si no tiene cuentas, entonces se agrega a la lista de personas_sin_cuentas y se salta
        if len(cuentas) == 0:
            personas_sin_cuentas.append(nomina.persona)
            continue

        # Tomar la cuenta de la persona que no tenga la clave 9, porque esa clave es la de DESPENSA
        su_cuenta = None
        for cuenta in cuentas:
            if cuenta.banco.clave != "9":
                su_cuenta = cuenta
                break

        # Si no tiene cuenta bancaria, entonces se agrega a la lista de personas_sin_cuentas y se salta
        if su_cuenta is None:
            personas_sin_cuentas.append(nomina.persona)
            continue

        # Tomar el banco de la cuenta de la persona
        su_banco = su_cuenta.banco

        # Incrementar la consecutivo del banco
        su_banco.consecutivo_generado += 1

        # Elaborar el numero de cheque, juntando la clave del banco y la consecutivo, siempre de 9 digitos
        num_cheque = f"{su_cuenta.banco.clave.zfill(2)}{su_banco.consecutivo_generado:07}"

        # Agregar la fila
        hoja.append(
            [
                nomina.quincena,
                nomina.centro_trabajo.clave,
                nomina.persona.rfc,
                nomina.persona.nombre_completo,
                nomina.persona.num_empleado,
                nomina.persona.modelo,
                nomina.plaza.clave,
                su_banco.nombre,
                su_banco.clave,
                su_cuenta.num_cuenta,
                nomina.importe,
                num_cheque,
            ]
        )

        # Incrementar contador
        contador += 1
        if contador % 100 == 0:
            click.echo(f"  Van {contador}...")

    # Actualizar los consecutivos de cada banco
    sesion.commit()

    # Determinar el nombre del archivo XLSX, juntando 'nominas' con la quincena y la fecha como YYYY-MM-DD HHMMSS
    nombre_archivo = f"nominas_{quincena}_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.xlsx"

    # Guardar el archivo XLSX
    libro.save(nombre_archivo)

    # Si hubo personas sin cuentas, entonces mostrarlas en pantalla
    if len(personas_sin_cuentas) > 0:
        click.echo("AVISO: Hubo personas sin cuentas:")
        for persona in personas_sin_cuentas:
            click.echo(f"- {persona.rfc} {persona.nombre_completo}")

    # Mensaje termino
    click.echo(f"Nominas terminado: {contador} nominas generadas en {nombre_archivo}")


@click.command()
@click.argument("quincena", type=str)
def generar_monederos(quincena: str):
    """Generar archivo XLSX con los monederos de una quincena"""

    # Validar quincena
    if re.match(QUINCENA_REGEXP, quincena) is None:
        click.echo("Quincena inválida")
        return

    # Iniciar sesion con la base de datos para que la alimentacion sea rapida
    sesion = database.session

    # Cargar solo el banco con la clave 9 que es PREVIVALE
    banco = Banco.query.filter_by(clave="9").first()
    if banco is None:
        click.echo("ERROR: No existe el banco con clave 9")
        return

    # Igualar el consecutivo_generado al consecutivo
    banco.consecutivo_generado = banco.consecutivo

    # Consultar las nominas de la quincena solo tipo DESPENSA
    nominas = Nomina.query.filter_by(quincena=quincena).filter_by(tipo="DESPENSA").filter_by(estatus="A").all()

    # Iniciar el archivo XLSX
    libro = Workbook()

    # Tomar la hoja del libro XLSX
    hoja = libro.active

    # Agregar la fila con las cabeceras de las columnas
    hoja.append(
        [
            "CT_CLASIF",
            "RFC",
            "TOT NET CHEQUE",
            "NUM CHEQUE",
            "NUM TARJETA",
            "QUINCENA",
            "MODELO",
        ]
    )

    # Bucle para crear cada fila del archivo XLSX
    contador = 0
    personas_sin_cuentas = []
    for nomina in nominas:
        # Tomar las cuentas de la persona
        cuentas = nomina.persona.cuentas

        # Si no tiene cuentas, entonces se agrega a la lista de personas_sin_cuentas y se salta
        if len(cuentas) == 0:
            personas_sin_cuentas.append(nomina.persona)
            continue

        # Tomar la cuenta de la persona que no tenga la clave 9, porque esa clave es la de DESPENSA
        su_cuenta = None
        for cuenta in cuentas:
            if cuenta.banco.clave == "9":
                su_cuenta = cuenta
                break

        # Si no tiene cuenta bancaria, entonces se agrega a la lista de personas_sin_cuentas y se salta
        if su_cuenta is None:
            personas_sin_cuentas.append(nomina.persona)
            continue

        # Incrementar el consecutivo_generado del banco
        banco.consecutivo_generado += 1

        # Elaborar el numero de cheque, juntando la clave del banco y el consecutivo_generado, siempre de 9 digitos
        num_cheque = f"{su_cuenta.banco.clave.zfill(2)}{banco.consecutivo_generado:07}"

        # Agregar la fila
        hoja.append(
            [
                "J",
                nomina.persona.rfc,
                nomina.importe,
                num_cheque,
                su_cuenta.num_cuenta,
                nomina.quincena,
                nomina.persona.modelo,
            ]
        )

        # Incrementar contador
        contador += 1
        if contador % 100 == 0:
            click.echo(f"  Van {contador}...")

    # Actualizar los consecutivo_generado de cada banco
    sesion.commit()

    # Determinar el nombre del archivo XLSX, juntando 'monederos' con la quincena y la fecha como YYYY-MM-DD HHMMSS
    nombre_archivo = f"monederos_{quincena}_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.xlsx"

    # Guardar el archivo XLSX
    libro.save(nombre_archivo)

    # Si hubo personas sin cuentas, entonces mostrarlas en pantalla
    if len(personas_sin_cuentas) > 0:
        click.echo("AVISO: Hubo personas sin cuentas:")
        for persona in personas_sin_cuentas:
            click.echo(f"- {persona.rfc} {persona.nombre_completo}")

    # Mensaje termino
    click.echo(f"Nominas terminado: {contador} monederos generadas en {nombre_archivo}")


@click.command()
@click.argument("quincena", type=str)
def generar_pensionados(quincena: str):
    """Generar archivo XLSX con los pensionados de una quincena"""

    # Validar quincena
    if re.match(QUINCENA_REGEXP, quincena) is None:
        click.echo("Quincena inválida.")
        return

    # Iniciar sesion con la base de datos para que la alimentacion sea rapida
    sesion = database.session

    # Consultar las nominas de la quincena, solo tipo SALARIO
    nominas = Nomina.query.filter_by(quincena=quincena).filter_by(tipo="SALARIO").filter_by(estatus="A").all()

    # Iniciar el archivo XLSX
    libro = Workbook()

    # Tomar la hoja del libro XLSX
    hoja = libro.active

    # Agregar la fila con las cabeceras de las columnas
    hoja.append(
        [
            "QUINCENA",
            "CENTRO DE TRABAJO",
            "RFC",
            "NOMBRE COMPLETO",
            "NUMERO DE EMPLEADO",
            "MODELO",
            "PLAZA",
            "NOMBRE DEL BANCO",
            "BANCO ADMINISTRADOR",
            "NUMERO DE CUENTA",
            "MONTO A DEPOSITAR",
            "NO DE CHEQUE",
        ]
    )

    # Bucle para crear cada fila del archivo XLSX
    contador = 0
    personas_sin_cuentas = []
    for nomina in nominas:
        # Si el modelo de la persona NO es 3, se omite
        if nomina.persona.modelo != 3:
            continue

        # Tomar las cuentas de la persona
        cuentas = nomina.persona.cuentas

        # Si no tiene cuentas, entonces se le crea una cuenta con el banco 10
        if len(cuentas) == 0:
            personas_sin_cuentas.append(nomina.persona)
            continue

        # Tomar la cuenta de la persona que no tenga la clave 9, porque esa clave es la de DESPENSA
        su_cuenta = None
        for cuenta in cuentas:
            if cuenta.banco.clave != "9":
                su_cuenta = cuenta
                break

        # Si no tiene cuenta bancaria, entonces se le crea una cuenta con el banco 10
        if su_cuenta is None:
            personas_sin_cuentas.append(nomina.persona)
            continue

        # Tomar el banco de la cuenta de la persona
        su_banco = su_cuenta.banco

        # Incrementar la consecutivo del banco
        su_banco.consecutivo_generado += 1

        # Elaborar el numero de cheque, juntando la clave del banco y la consecutivo, siempre de 9 digitos
        num_cheque = f"{su_cuenta.banco.clave.zfill(2)}{su_banco.consecutivo_generado:07}"

        # Agregar la fila
        hoja.append(
            [
                nomina.quincena,
                nomina.centro_trabajo.clave,
                nomina.persona.rfc,
                nomina.persona.nombre_completo,
                nomina.persona.num_empleado,
                nomina.persona.modelo,
                nomina.plaza.clave,
                su_banco.nombre,
                su_banco.clave,
                su_cuenta.num_cuenta,
                nomina.importe,
                num_cheque,
            ]
        )

        # Incrementar contador
        contador += 1
        if contador % 100 == 0:
            click.echo(f"  Van {contador}...")

    # Actualizar los consecutivos de cada banco
    sesion.commit()

    # Determinar el nombre del archivo XLSX, juntando 'nominas' con la quincena y la fecha como YYYY-MM-DD HHMMSS
    nombre_archivo = f"pensionados_{quincena}_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.xlsx"

    # Guardar el archivo XLSX
    libro.save(nombre_archivo)

    # Si hubo personas sin cuentas, entonces mostrarlas en pantalla
    if len(personas_sin_cuentas) > 0:
        click.echo("AVISO: Hubo personas sin cuentas:")
        for persona in personas_sin_cuentas:
            click.echo(f"- {persona.rfc} {persona.nombre_completo}")

    # Mensaje termino
    click.echo(f"Nominas terminado: {contador} pensionados generados en {nombre_archivo}")


@click.command()
@click.argument("quincena", type=str)
def generar_dispersiones_pensionados(quincena: str):
    """Generar archivo XLSX con las dispersiones pensionados de una quincena"""

    # Validar quincena
    if re.match(QUINCENA_REGEXP, quincena) is None:
        click.echo("Quincena inválida.")
        return

    # TODO: Validar que la quincena este abierta

    # Iniciar sesion con la base de datos para que la alimentacion sea rapida
    sesion = database.session

    # Consultar las nominas de la quincena, solo tipo SALARIO
    nominas = Nomina.query.filter_by(quincena=quincena).filter_by(tipo="SALARIO").filter_by(estatus="A").all()

    # Iniciar el archivo XLSX
    libro = Workbook()

    # Tomar la hoja del libro XLSX
    hoja = libro.active

    # Agregar la fila con las cabeceras de las columnas
    hoja.append(
        [
            "CONSECUTIVO",
            "FORMA DE PAGO",
            "TIPO DE CUENTA",
            "BANCO RECEPTOR",
            "CUENTA ABONO",
            "IMPORTE PAGO",
            "CLAVE BENEFICIARIO",
            "RFC",
            "NOMBRE",
            "REFERENCIA PAGO",
            "CONCEPTO PAGO",
        ]
    )

    # Bucle para crear cada fila del archivo XLSX
    contador = 0
    personas_sin_cuentas = []
    for nomina in nominas:
        # Si el modelo de la persona NO es 3, se omite
        if nomina.persona.modelo != 3:
            continue

        # Tomar las cuentas de la persona
        cuentas = nomina.persona.cuentas

        # Si no tiene cuentas, entonces se le crea una cuenta con el banco 10
        if len(cuentas) == 0:
            personas_sin_cuentas.append(nomina.persona)
            continue

        # Tomar la cuenta de la persona que no tenga la clave 9, porque esa clave es la de DESPENSA
        su_cuenta = None
        for cuenta in cuentas:
            if cuenta.banco.clave != "9":
                su_cuenta = cuenta
                break

        # Si no tiene cuenta bancaria, entonces se le crea una cuenta con el banco 10
        if su_cuenta is None:
            personas_sin_cuentas.append(nomina.persona)
            continue

        # Definir referencia_pago, se forma con los dos ultimos caracteres y los caracteres tercero y cuarto de la quincena
        referencia_pago = f"{quincena[-2:]}{quincena[2:4]}"

        # Definir concepto_pago, se forma con el texto "QUINCENA {dos digitos} PENSIONADOS"
        concepto_pago = f"QUINCENA {quincena[-2:]} PENSIONADOS"

        # Agregar la fila
        hoja.append(
            [
                contador + 1,
                "04",
                "9",
                su_cuenta.banco.clave_dispersion_pensionados,
                su_cuenta.num_cuenta,
                nomina.importe,
                contador + 1,
                nomina.persona.rfc,
                nomina.persona.nombre_completo,
                referencia_pago,
                concepto_pago,
            ]
        )

        # Incrementar contador
        contador += 1
        if contador % 100 == 0:
            click.echo(f"  Van {contador}...")

    # Determinar el nombre del archivo XLSX, juntando 'nominas' con la quincena y la fecha como YYYY-MM-DD HHMMSS
    nombre_archivo = f"dispersiones_pensionados_{quincena}_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.xlsx"

    # Guardar el archivo XLSX
    libro.save(nombre_archivo)

    # Si hubo personas sin cuentas, entonces mostrarlas en pantalla
    if len(personas_sin_cuentas) > 0:
        click.echo("AVISO: Hubo personas sin cuentas:")
        for persona in personas_sin_cuentas:
            click.echo(f"- {persona.rfc} {persona.nombre_completo}")

    # Mensaje termino
    click.echo(f"Nominas terminado: {contador} dispersiones pensionados generados en {nombre_archivo}")


cli.add_command(alimentar)
cli.add_command(generar_nominas)
cli.add_command(generar_monederos)
cli.add_command(generar_pensionados)
cli.add_command(generar_dispersiones_pensionados)